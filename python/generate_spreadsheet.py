"""
Excel Spreadsheet Generator for Piston Kinematics & Dynamics
Author: Rahul Parmar
Description:
    Generates a professionally styled Microsoft Excel (.xlsx) workbook
    containing full slider-crank mathematical calculations, parameter inputs,
    summary KPIs, degree-by-degree data table, and embedded Excel charts.
"""

import sys
import os
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, Series

from engine_kinematics import SliderCrankKinematics


def generate_piston_spreadsheet(
    bore: float = 85.0,
    stroke: float = 88.0,
    con_rod_length: float = 145.0,
    rpm: float = 6000.0,
    reciprocating_mass: float = 0.45,
    angle_step: float = 2.0,
    output_filepath: str = "piston_kinematics_sheet.xlsx"
) -> str:
    """
    Generate a full-featured Excel workbook.
    """
    calc = SliderCrankKinematics(
        bore=bore,
        stroke=stroke,
        con_rod_length=con_rod_length,
        rpm=rpm,
        reciprocating_mass=reciprocating_mass
    )
    summary = calc.get_summary_statistics()
    cycle_data = calc.compute_cycle(angle_step=angle_step, max_angle=360.0)

    wb = Workbook()
    
    # -------------------------------------------------------------
    # 1. Styles Definition
    # -------------------------------------------------------------
    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800

    section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    section_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Slate 200

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid") # Blue 600
    header_sub_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid") # Sky 600

    kpi_title_font = Font(name="Calibri", size=9, bold=False, color="64748B")
    kpi_val_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
    kpi_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    data_font = Font(name="Calibri", size=10, color="000000")
    data_alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # -------------------------------------------------------------
    # Sheet 1: Dashboard & Summary
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Engine Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    # Title Banner
    ws_summary.merge_cells("A1:G2")
    title_cell = ws_summary["A1"]
    title_cell.value = "PISTON KINEMATICS & DYNAMICS ANALYSIS"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_summary["A3"] = "Project by: Rahul Parmar | Advanced Slider-Crank Analysis"
    ws_summary["A3"].font = Font(name="Calibri", size=10, italic=True, color="64748B")

    # Inputs Block
    ws_summary["A5"] = "1. ENGINE DESIGN PARAMETERS (INPUTS)"
    ws_summary["A5"].font = section_font
    ws_summary["A5"].fill = section_fill
    ws_summary.merge_cells("A5:D5")

    inputs_meta = [
        ("Cylinder Bore (D)", bore, "mm"),
        ("Piston Stroke (S)", stroke, "mm"),
        ("Connecting Rod Length (L)", con_rod_length, "mm"),
        ("Engine Rotational Speed (N)", rpm, "RPM"),
        ("Reciprocating Mass (m_r)", reciprocating_mass, "kg"),
        ("Crank Radius (r = S/2)", stroke / 2.0, "mm"),
        ("Rod-to-Crank Ratio (n = L/r)", con_rod_length / (stroke / 2.0), "ratio"),
        ("Crank-to-Rod Ratio (lambda = r/L)", (stroke / 2.0) / con_rod_length, "ratio"),
    ]

    row = 6
    for label, val, unit in inputs_meta:
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        val_cell = ws_summary.cell(row=row, column=2, value=val)
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = "0.00" if isinstance(val, float) else "0"
        ws_summary.cell(row=row, column=3, value=unit).font = Font(color="64748B")
        for col in range(1, 4):
            ws_summary.cell(row=row, column=col).border = thin_border
        row += 1

    # Key Performance Indicators (KPIs)
    ws_summary["E5"] = "2. CRITICAL PERFORMANCE METRICS"
    ws_summary["E5"].font = section_font
    ws_summary["E5"].fill = section_fill
    ws_summary.merge_cells("E5:H5")

    kpi_meta = [
        ("Engine Displacement", summary["displacement_volume_cc"], "cc / cm3"),
        ("Angular Velocity (omega)", summary["angular_velocity_rad_s"], "rad/s"),
        ("Mean Piston Speed (v_mean)", summary["mean_piston_speed_mps"], "m/s"),
        ("Max Piston Velocity", summary["max_velocity_mps"], "m/s"),
        ("Crank Angle @ Max Velocity", summary["angle_at_max_velocity_deg"], "deg (from TDC)"),
        ("Max Acceleration (TDC)", summary["accel_at_tdc_mps2"], "m/s^2"),
        ("Max Acceleration (TDC in g's)", summary["accel_at_tdc_g"], "g"),
        ("Acceleration @ BDC", summary["accel_at_bdc_mps2"], "m/s^2"),
        ("Max Reciprocating Inertia Force (TDC)", abs(summary["inertia_force_at_tdc_N"]), "N"),
        ("Inertia Force @ TDC (kN)", abs(summary["inertia_force_at_tdc_kN"]), "kN"),
        ("Inertia Force @ BDC (kN)", abs(summary["inertia_force_at_bdc_kN"]), "kN"),
    ]

    row = 6
    for label, val, unit in kpi_meta:
        ws_summary.cell(row=row, column=5, value=label).font = Font(bold=True)
        val_cell = ws_summary.cell(row=row, column=6, value=val)
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = "0.00"
        ws_summary.cell(row=row, column=7, value=unit).font = Font(color="64748B")
        for col in range(5, 8):
            ws_summary.cell(row=row, column=col).border = thin_border
        row += 1

    # Auto-adjust column widths for summary sheet
    for col in ws_summary.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 3, 14)

    # -------------------------------------------------------------
    # Sheet 2: Detailed Kinematics Data Table (Spreadsheet Data)
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Kinematics Data")
    ws_data.views.sheetView[0].showGridLines = True

    # Header Row
    headers = [
        "Crank Angle θ (°)",
        "Crank Angle θ (rad)",
        "Con-Rod Angle β (°)",
        "Displacement x (mm)",
        "Velocity v (m/s)",
        "Accel Exact a (m/s²)",
        "Accel Primary (m/s²)",
        "Accel Secondary (m/s²)",
        "Inertia Force Fi (N)",
        "Inertia Force Fi (kN)",
    ]

    for col_idx, h_text in enumerate(headers, start=1):
        cell = ws_data.cell(row=1, column=col_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws_data.row_dimensions[1].height = 28

    # Populate degree-by-degree data
    num_points = len(cycle_data["crank_angle_deg"])
    for i in range(num_points):
        row_num = i + 2
        theta_deg = float(cycle_data["crank_angle_deg"][i])
        theta_rad = (theta_deg * 3.141592653589793) / 180.0
        beta_deg = float(cycle_data["conrod_angle_deg"][i])
        disp_mm = float(cycle_data["displacement_mm"][i])
        vel_mps = float(cycle_data["velocity_mps"][i])
        acc_mps2 = float(cycle_data["acceleration_mps2"][i])
        acc_p_mps2 = float(cycle_data["acceleration_primary_mps2"][i])
        acc_s_mps2 = float(cycle_data["acceleration_secondary_mps2"][i])
        inertia_N = float(cycle_data["inertia_force_N"][i])
        inertia_kN = inertia_N / 1000.0

        row_vals = [
            theta_deg,
            theta_rad,
            beta_deg,
            disp_mm,
            vel_mps,
            acc_mps2,
            acc_p_mps2,
            acc_s_mps2,
            inertia_N,
            inertia_kN
        ]

        is_even = (i % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_data.cell(row=row_num, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="right")
            if is_even:
                cell.fill = data_alt_fill
            
            # Formatting
            if col_idx in [1]:
                cell.number_format = "0.0"
            elif col_idx in [2, 3]:
                cell.number_format = "0.000"
            elif col_idx in [4, 5]:
                cell.number_format = "0.00"
            elif col_idx in [6, 7, 8]:
                cell.number_format = "0.0"
            elif col_idx in [9]:
                cell.number_format = "#,##0.0"
            elif col_idx in [10]:
                cell.number_format = "0.000"

    # Auto-adjust column widths for data sheet
    for col in ws_data.columns:
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = 16

    # -------------------------------------------------------------
    # Add Embedded Excel Charts
    # -------------------------------------------------------------
    # Chart 1: Velocity vs Crank Angle
    chart_vel = LineChart()
    chart_vel.title = "Piston Velocity vs Crank Angle (θ)"
    chart_vel.style = 13
    chart_vel.y_axis.title = "Velocity (m/s)"
    chart_vel.x_axis.title = "Crank Angle (°)"
    chart_vel.width = 16
    chart_vel.height = 10

    data_ref_v = Reference(ws_data, min_col=5, min_row=1, max_row=num_points + 1)
    cats_ref = Reference(ws_data, min_col=1, min_row=2, max_row=num_points + 1)
    chart_vel.add_data(data_ref_v, titles_from_data=True)
    chart_vel.set_categories(cats_ref)
    ws_data.add_chart(chart_vel, "L2")

    # Chart 2: Acceleration vs Crank Angle
    chart_acc = LineChart()
    chart_acc.title = "Piston Acceleration vs Crank Angle (θ)"
    chart_acc.style = 10
    chart_acc.y_axis.title = "Acceleration (m/s²)"
    chart_acc.x_axis.title = "Crank Angle (°)"
    chart_acc.width = 16
    chart_acc.height = 10

    data_ref_a = Reference(ws_data, min_col=6, max_col=8, min_row=1, max_row=num_points + 1)
    chart_acc.add_data(data_ref_a, titles_from_data=True)
    chart_acc.set_categories(cats_ref)
    ws_data.add_chart(chart_acc, "L18")

    # Chart 3: Reciprocating Inertia Force vs Crank Angle
    chart_force = LineChart()
    chart_force.title = "Reciprocating Inertia Force vs Crank Angle (θ)"
    chart_force.style = 14
    chart_force.y_axis.title = "Inertia Force (kN)"
    chart_force.x_axis.title = "Crank Angle (°)"
    chart_force.width = 16
    chart_force.height = 10

    data_ref_f = Reference(ws_data, min_col=10, min_row=1, max_row=num_points + 1)
    chart_force.add_data(data_ref_f, titles_from_data=True)
    chart_force.set_categories(cats_ref)
    ws_data.add_chart(chart_force, "L34")

    # Save workbook
    wb.save(output_filepath)
    print(f"[OK] Spreadsheet generated successfully at: {os.path.abspath(output_filepath)}")
    return os.path.abspath(output_filepath)


if __name__ == "__main__":
    generate_piston_spreadsheet(output_filepath="piston_kinematics_sheet.xlsx")
